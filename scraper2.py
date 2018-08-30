import requests, re, html2text, time, random
from bs4 import BeautifulSoup
from openpyxl import load_workbook,Workbook
from collections import Counter

workbook = 'CompanyProfilespja_MASTER.xlsx'

wb = load_workbook(workbook)
ws = wb[wb.sheetnames[2]]
startRow = 1500
endRow = 1900
endRow += 1
scrubbedBy = "Khristan"
successfulEmails = 0
successfulPNs = 0
numSearches = 3
searchEngine = "Bing"
for row in range(startRow,endRow):
    print("Current Row: "+str(row))
    prevScrubber = ws['B'+str(row)].value
    chosenWebsite = ws['U'+str(row)].value
    if(chosenWebsite):
        validUrl = re.search('((?:www.)?([a-zA-Z0-9]+(?:-[a-zA-Z0-9]+)*(?:\.[a-zA-Z]+)+))',chosenWebsite,re.M|re.I)

        if  validUrl and (prevScrubber == scrubbedBy or not prevScrubber) :
            chosenEmail = ws['O'+str(row)].value
            chosenPN = ws['B'+str(row)].value
            foundEmails = []
            foundPNs = []
            html = ""
            htmlText = ""
            try:
                html += requests.get(chosenWebsite,{'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:54.0) Gecko/20100101 Firefox/54.0'})
            except:
                print("Direct Link Broken :(")


            text_maker = html2text.HTML2Text()
            text_maker.ignore_emphasis = True
            htmlText = text_maker.handle(BeautifulSoup(html.content,"html5lib").prettify())
            foundEmails += re.findall(r'([\-\w\d.]+\s*\@\s*'+re.escape(searchQuery.group(1))+')',htmlText)
            foundPNs += re.findall(r'(\d{3}[-\.\s]??\d{3}[-\.\s]??\d{4}|\(\d{3}\)\s*\d{3}[-\.\s]??\d{4}|\d{3}[-\.\s]??\d{4})',htmlText)

print('Found E-mails: '+ str(100*(successfulEmails/abs(endRow-startRow)))+'%')
print('# of Emails: '+ str(successfulEmails))
print('Found Phone#s: '+ str(100*(successfulPNs/abs(endRow-startRow)))+'%')
print('# of Emails: '+ str(successfulPNs))
