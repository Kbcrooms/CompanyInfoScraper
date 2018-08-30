import requests, re, html2text, time, random
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
from collections import Counter
wb = load_workbook('CompanyProfilespja_MASTER.xlsx')
ws = wb[wb.sheetnames[2]]

cwb = load_workbook('import.xlsx')
cws = cwb[cwb.sheetnames[0]]
startRow = 2300
endRow = 3000
scrubbedName = "Khristan"
iCurs = 3
for row in range(startRow,endRow):
    prevScrubber = ws['B'+str(row)].value
    website = ws['U'+str(row)].value
    email = ws['O'+str(row)].value
    phone = ws['Q'+str(row)].value
    company = ws['I'+str(row)].value
    imported = ws['C'+str(row)].value
    if(prevScrubber == "Khristan" and email and not imported):
        iCurs = iCurs +1
        cws['F'+str(iCurs)].value = company
        cws['G'+str(iCurs)].value = company
        cws['J'+str(iCurs)].value = "No"
        cws['M'+str(iCurs)].value = email
        cws['O'+str(iCurs)].value = phone
        ws['C'+str(row)].value = "Khristan"

cwb.save('import.xlsx')
wb.save('CompanyProfilespja_MASTER.xlsx')
