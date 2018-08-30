import requests, re, html2text, time, random
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
from collections import Counter
wb = load_workbook('CompanyProfilespja_MASTER.xlsx')
ws = wb[wb.sheetnames[2]]
startRow = 2500
endRow = 3000
scrubbedName = "Khristan"
emailCounter = 0
for row in range(startRow,endRow):
    print(row)
    prevScrubber = ws['B'+str(row)].value
    website = ws['U'+str(row)].value
    email = ws['O'+str(row)].value
    #website = 'sysplus.com'
    if  (prevScrubber == scrubbedName or not prevScrubber or (prevScrubber=="Phyllis" and not email)) and website and not email:

        foundEmails = []
        foundPhoneNumbers = []
        chosenEmail = ""
        chosenPN = ""
        searchQuery = re.search('(?:www.)?([a-zA-Z0-9]+(?:-[a-zA-Z0-9]+)*(?:\.[a-zA-Z]+)+)',website,re.M|re.I)
        if searchQuery:
            searchEmail = "info%40"+searchQuery.group(1)
            resultCounter = 0
            for i in range(0,3):
                #time.sleep((random.random()*7)+3)
                #Google
                #searchPage = '&start='+ str(resultCounter)
                #Bing
                searchPage = '&first='+ str(resultCounter+1)
                if i == 0:
                    resultCounter += 8
                else:
                    resultCounter +=13
                searchURL = ""
                if i==0:
                    #Bing
                    searchURL = "https://bing.com/search?" + "q="+ searchQuery.group(1)+'%20contact%20email%20phone' +searchPage
                    print(searchURL)
                    #Google
                    #searchURL = "https://google.com/search?source=hp&" + "biw=1920&bih=989" +"&q="+ searchEmail +"&oq="+searchEmail + searchPage +'&sourceid=firefox&ie=UTF-8'
                elif i>0:
                    #Bing
                    searchURL = "https://bing.com/search?" + "q="+ searchQuery.group(1)+'%20contact%20email%20phone' +searchPage
                    print(searchURL)
                    #Google
                    #searchURL = "https://google.com/search?q=" + searchEmail +"&oq="+searchEmail + searchPage +'&sa=N&biw=1920&bih=989'
                html = requests.get(searchURL,{'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:54.0) Gecko/20100101 Firefox/54.0'})
                text_maker = html2text.HTML2Text()
                text_maker.ignore_emphasis = True

                htmlText = text_maker.handle(BeautifulSoup(html.content,"html5lib").prettify())
                #print(htmlText)
                foundEmails += re.findall(r'([\-\w\d.]+\s*\@\s*'+re.escape(searchQuery.group(1))+')',htmlText)
                foundPhoneNumbers += re.findall(r'(\d{3}[-\.\s]??\d{3}[-\.\s]??\d{4}|\(\d{3}\)\s*\d{3}[-\.\s]??\d{4}|\d{3}[-\.\s]??\d{4})',htmlText)
                #print("CONTENT START")
                #print(htmlText)
                #print (htmlText)
                #print("CONTENT END")
            for i in range(0,len(foundEmails)):
                foundEmails[i]="".join(foundEmails[i].lower().split())
            commonEmails = Counter(foundEmails).most_common(2)
            eLen = len(commonEmails)
            if(eLen>0):
                emailCounter+=1
                if(eLen>1 and commonEmails[1][1]>=3):
                    chosenEmail = commonEmails[1][0]
                else:
                    chosenEmail = commonEmails[0][0]
            print(chosenEmail)
            for i in range(0,len(foundPhoneNumbers)):
                foundPhoneNumbers[i]=re.sub("[^0-9]","", foundPhoneNumbers[i])
            commonPN = Counter(foundPhoneNumbers).most_common(len(foundPhoneNumbers))
            pnLen = len(commonPN)
            for pn in commonPN:
                if(len(pn[0])==10 and int(pn[0][:1]) >= 2):
                    chosenPN = pn[0][:3] + "-" + pn[0][3:6] + "-" + pn[0][6:]
                    break
            print(chosenPN)
            chosenPN = chosenPN.strip()
            chosenEmail = chosenEmail.strip()
            if(len(chosenEmail.strip())>0):
                emailCounter+=1
                ws['O'+str(row)] = chosenEmail
            if(len(chosenPN.strip())>0):
                ws['Q'+str(row)] = chosenPN
            if len(chosenPN.strip())>0 or len(chosenEmail.strip())>0:
                ws['B'+str(row)] = scrubbedName
                print("Added Name")
wb.save('CompanyProfilespja_MASTER.xlsx')
print(emailCounter)
